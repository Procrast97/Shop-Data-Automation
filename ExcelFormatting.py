from email import encoders
from email.mime.multipart import MIMEMultipart
from datetime import datetime as dt
import pandas as pd
from io import BytesIO
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import os
from openpyxl import load_workbook
import smtplib
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
import openpyxl
from openpyxl.utils import get_column_letter
import matplotlib.font_manager as fm
import matplotlib as mpl

from DataExtract import EMAIL

FILE_PATH = "Sales_Data.xlsx"
PDF_PATH = "Data_PDF.pdf"

fm.fontManager.addfont(r'C:\Users\user\OneDrive\Documents\Work\HKwork\Coding\Vincent\Project1-SalesScrapper\Fonts\NotoSansHK.ttf')
plt.rcParams['font.sans-serif'] = ['Noto Sans HK', 'NotoSansHK', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


class ExcelCreation:

    def __init__(self):
        self.connection = smtplib.SMTP_SSL('smtp.gmail.com')

    def create_excelsheet(self, data):
        # EXCEPTION CATCH - Keep getting a "data" not defined error:
        try:
            formated_data = pd.DataFrame(data)
            horizontal_form = formated_data.T
        except NameError:
            print("A 'NameError' has occurred, trying again...")
            formated_data = pd.DataFrame(data)
            horizontal_form = formated_data.T

            # ## Selecting the shop names for saving the data:
            # shop_name = SHOPS[loops][shops - 1]

        ## Sort data into a CSV file (excel):
        formated_data = pd.DataFrame(horizontal_form)

        ## Check if the file exists:
        if os.path.exists(FILE_PATH):
            # book = load_workbook(filename=FILE_PATH)

            wb = openpyxl.Workbook()

            wb.save(FILE_PATH)

            with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
                formated_data.to_excel(writer, sheet_name="Sheet", index=False, header=True)

        ## If the file_path doesn't exist:
        else:
            with pd.ExcelWriter(FILE_PATH, engine="openpyxl") as writer:
                formated_data.to_excel(writer, index=False)

        #        if formated_data is None or formated_data.empty:
        #            pass
        #        else:
        #            removed_colms = formated_data.drop(columns=["Date", "Customer"])

        self.auto_fit_excel_columns(FILE_PATH)
        return formated_data

    def removing_columns(self, formated_data):
        colm_removed = formated_data.drop(columns=["Date", "Customer"])
        return colm_removed

    def create_pdf(self, pd_data):

        pdf_buffer = BytesIO()

        # Create the figure and table
        fig, ax = plt.subplots(figsize=(12, 4))
        ax.axis('tight')
        ax.axis('off')
        the_table = ax.table(
            cellText=pd_data.values,
            colLabels=pd_data.columns,
            cellLoc='center',
            loc='center'
        )

        the_table.auto_set_font_size(False)
        the_table.set_fontsize(10)
        the_table.scale(1.2, 1.5)

        plt.title("Data_PDF.pdf")

        # Save to PdfPages instead of to a file
        with PdfPages(pdf_buffer) as pdf:
            pdf.savefig(fig, bbox_inches='tight')  # Add the figure to PdfPages

        plt.close(fig)  # Close the figure
        pdf_buffer.seek(0)

        if len(pdf_buffer.getvalue()) == 0:
            raise ValueError("PDF generation failed - empty buffer.")
        else:
            return pdf_buffer

    def auto_fit_excel_columns(self, file_path):
        work_book = load_workbook(file_path)
        work_sheet = work_book.active

        for column in work_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjust_width = min(max_length + 2, 50)

            work_sheet.column_dimensions[column_letter].width = adjust_width

        work_book.save(file_path)

    def send_mail(self, email, password, empty_shop_list, sales_shops_list, total, r_data, send_condition):

        if not send_condition:
            self.connection.login(email, password)
            self.connection.sendmail(from_addr=email,
                                     # to_addrs=os.environ.get('ACCOUNT_EMAIL'),
                                     to_addrs="zanderhenning@gmail.com",
                                     msg="Subject: No Sales\n\n No sales were recorded during today's scan.")
        else:

            empty_shops = ""
            sales_shops = ""

            for a in sales_shops_list:
                sales_shops = sales_shops + a + ", "

            for a in empty_shop_list:
                empty_shops = empty_shops + a + ", "

            message = MIMEMultipart()
            message['From'] = email
            message['To'] = "zanderhenning@gmail.com"
            message['Subject'] = f"Sales Data - {dt.now().strftime('%m/%d/%Y')}, Total: ${total}"

            message.attach(
                MIMEText(f"The following Locations (Store) has yielded no sales: {empty_shops}.\nThe following Locations (Store) "
                         f"yielded sales: {sales_shops}\nSales Data: \n{r_data}\n\nPlease refer to the attached Excel/PDF Document attached for "
                         f"more details on current sales.\n\nRegards\nVector Digital"))

            file = os.path.basename(FILE_PATH)
            # pdf_file = os.path.basename(PDF_PATH)
            pdf_buffer = self.create_pdf(r_data)

            with open(FILE_PATH, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

            encoders.encode_base64(part)

            part.add_header("Content-Disposition", f"attachment; filename={file}")

            if pdf_buffer is None or pdf_buffer.getvalue() == b'':
                message.attach(r_data)
                # self.connection.login(email, password)
                # self.connection.sendmail(from_addr=email,
                #                      to_addrs="zanderhenning@gmail.com",
            # 	 msg='Subject: PDF CREATION ERROR\n\nThe pdf creation code failed due to NO DATA!'
            #                 )
            else:
                part_pdf = MIMEBase("application", "pdf")
                part_pdf.set_payload(pdf_buffer.getvalue())
                encoders.encode_base64(part_pdf)
                part_pdf.add_header(
                    "Content-Disposition",
                    "attachment; filename=Sales_Report.pdf"
                )

                message.attach(part_pdf)

            message.attach(part)

            text = message.as_string()

            self.connection.login(email, password)
            self.connection.sendmail(from_addr=email,
                                     to_addrs="vincent.2998@yahoo.com",
                                     msg=text,
                                     )
            # self.connection.sendmail(from_addr=email,
            #                          to_addrs="zanderhenning@gmail.com",
            #                          # msg='Subject: No problems\n\nThe process is completed. No errors occured.'
            #                          msg=text,
            #                          )




