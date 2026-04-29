#For Excel Exporting:
from idlelib.browser import file_open

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import session, render_template, redirect, url_for, Flask, send_file
from DBModels import ItemData, ShopLoc
from datetime import datetime
from io import BytesIO

#For PDF exporting:
## I Learned this from Claude AI, it taught me how to install fonts locally in your project files so that you do not have
# to continuously download certain modules and classes to make it work.
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER



class ExcelExport:
    def __init__(self, report_data=None):
        self.report_data = report_data


    def build_item_query(self, selected_shop_id, selected_location_id):
        """
        Building an ItemData query based on selected shop_id and location_id.
        This will return filtered query object for all 3 processes.
        """
        query = ItemData.query
        if selected_shop_id == 'all' and selected_location_id == 'all':
            pass

        elif selected_shop_id != 'all' and selected_location_id != 'all':
            query = query.filter_by(loc_id=int(selected_location_id))

        elif selected_shop_id != 'all' and selected_location_id == 'all':
            #This is for a specific shop with all it's corresponding locations.
            shop_locations = ShopLoc.query.filter_by(shop_id=int(selected_shop_id)).all()
            location_ids = [loc.id for loc in shop_locations]
            query = query.filter(ItemData.loc_id.in_(location_ids))

        elif selected_shop_id == 'all' and selected_location_id != 'all':
            query = query.filter_by(loc_id=int(selected_location_id))

        return query


    def export_excel(self):
        if not self.report_data:
            return redirect(url_for('generate_report'))

        #Build the query if the data is given:
        query = ItemData.query

        selected_shop_id = self.report_data['shop_id']
        selected_location_id = self.report_data['loc_id']
        date_from = datetime.strptime(self.report_data['date_from'], '%d/%m/%Y')
        date_to = datetime.strptime(self.report_data['date_to'], '%d/%m/%Y')

        queried_data = self.build_item_query(selected_shop_id, selected_location_id)

        formated_from = date_from.strftime('%d/%m/%Y')
        formated_to = date_to.strftime('%d/%m/%Y')
        queried_data = queried_data.filter(ItemData.date.between(formated_from, formated_to)).all()

        #Still not very confident with Excel sheets - Used AI for guidance:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        #Now we add the created title, commands are like clicking the on Excel commands themselves:
        ws.merge_cells("A1:H1")
        ws['A1'] = 'Sales Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        #Adding Date Range:
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Period: {date_from.strftime('%d/%m/%Y')} to {date_to.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        #Next, we add the headers:
        headers = ['Date', 'Shop Name', 'Customer', 'Description', 'Product ID', 'Qty', 'Price', 'Subtotal']
        ws.append([]) #This is to add an empty row first.
        ws.append(headers)

        #Styling the headers:
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        #Then, we add the data.
        total_qty = 0
        total_revenue = 0.00

        for item in queried_data:
            ws.append([
                item.date,
                item.shop_name,
                item.customer,
                item.description,
                item.product_id,
                item.qty,
                item.price,
                item.subtotal
            ])
            total_qty += item.qty
            total_revenue += float(item.subtotal.split(" ")[1])


        #Add totals row:
        totals_row = ws.max_row + 1
        ws.append(['', '', '', '', 'TOTALS:', total_qty, '', f"$ {total_revenue}"])

        #Styling the totals row:
        for cell in ws[totals_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill()

        #Adjust Column Widths:
        for column in ws.columns:
            max_width = 0
            column_letter = None

            for cell in column:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue

                if column_letter is None:
                    column_letter = cell.column_letter

                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_width:
                            max_width = cell_length

                except:
                    pass

            #Just add a little extra padding:
            if max_width > 0:
                adjusted_width = max(10, min(max_width + 4, 50))
                ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        #Files response
        filename = f"sales_report_{date_from.strftime('%Y%m%d')}_TO_{date_to.strftime('%Y%m%d')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )


    def export_pdf(self):
        """
        This will all be done with the assistance of Claude AI,
        as I am not yet very familiar with how to export PDF Files.
        """
        if not self.report_data:
            return redirect(url_for('generate_report'))

        pdfmetrics.registerFont(TTFont("NotoSansHK", "./Fonts/NotoSansHK.ttf"))
        pdfmetrics.registerFont(TTFont("NotoSansHK-Bold", "./Fonts/NotoSansHK-Bold.ttf"))

        query = ItemData.query

        selected_shop_id = self.report_data['shop_id']
        selected_location_id = self.report_data['loc_id']
        date_from = datetime.strptime(self.report_data['date_from'], '%d/%m/%Y')
        date_to = datetime.strptime(self.report_data['date_to'], '%d/%m/%Y')

        report_data = self.build_item_query(selected_shop_id, selected_location_id)

        formated_from = date_from.strftime('%d/%m/%Y')
        formated_to = date_to.strftime('%d/%m/%Y')
        report_data = report_data.filter(ItemData.date.between(formated_from, formated_to)).all()

        #First we need to create PDF in memory:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []

        #NOW WE PHYSICALLY CODE/BUILD OUR PDF:

        #Now styling:
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Stock Report Sheet",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        # The title:
        title = Paragraph("Sales Report", title_style)
        elements.append(title)

        #Date Range:
        date_range = Paragraph(
            f"Period: {date_from.strftime('%d/%m/%Y')} to {date_to.strftime('%d/%m/%Y')}",
            styles['Normal']
        )
        elements.append(date_range)
        elements.append(Spacer(1, 20))

        #Table Data
        data = [['Date', 'Shop', 'Customer', 'Description', 'Product ID', 'Qty', 'Price', 'Subtotal']]

        total_qty = 0
        total_revenue = 0

        for item in report_data:
            price = float(item.price.strip("$ ").strip("-"))
            subtotal = float(item.subtotal.strip("$ ").strip("-"))
            data.append([
                item.date,
                item.shop_name,
                item.customer,
                item.description[:30], #Shorten if too long
                item.product_id,
                str(item.qty),
                f'${price:.2f}',
                f'${subtotal:.2f}',
            ])

            total_qty += item.qty
            total_revenue += float(item.subtotal.split(" ")[1])

        #Adding the totals row:
        data.append(['', '', '', '', 'TOTALS:', str(total_qty), '', f'$ {total_revenue:.2f}'])

        #Create the Table:
        table = Table(data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 2*inch, 1*inch, 0.6*inch, 0.8*inch, 0.8*inch])

        #Then table styling:
        table.setStyle(TableStyle([

            #Header styling:
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'NotoSansHK-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            #Data Rows styling:
            ('FONTNAME', (0, 1), (-1, -2), 'NotoSansHK'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),

            #Totals rows styling:
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D3D3D3')),
            ('FONTNAME', (0, -1), (-1, -1), 'NotoSansHK-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
        ]))

        elements.append(table)

        #Now we build the PDF with what we added to the elements list above:
        doc.build(elements)
        buffer.seek(0)

        #Now the response:
        filename = f"sales_report_{date_from.strftime('%d/%m/%Y')}_TO_{date_to.strftime('%d/%m/%Y')}.pdf"

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

